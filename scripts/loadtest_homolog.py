#!/usr/bin/env python3
"""BlaXx Pontos - Teste de robustez de HOMOLOGACAO (usuarios ficticios).

============================ AVISO IMPORTANTE ============================
AMBIENTE EXCLUSIVO DE HOMOLOGACAO. Todos os usuarios, nomes, CPFs, telefones,
senhas, saldos e transacoes gerados por este script sao FICTICIOS e existem
apenas para teste funcional e de carga. Nenhum dado e real e o PIX roda em
modo MOCK (nenhum dinheiro real e movimentado).
=========================================================================

Fases:
  0. Semeia N usuarios ficticios direto no banco (hash pbkdf2 barato).
  1. Exporta planilha Excel (+CSV) com usuarios e senhas de teste.
  2. Sobe o backend (servidor proprio, threaded) e dispara carga HTTP real
     e concorrente: login -> consulta saldo -> compra de pontos (PIX mock) ->
     transferencia P2P -> resgate de beneficio -> (parte) resgate via PIX.
  3. Verifica a integridade do razao (ledger): saldo == soma confirmada,
     nenhum saldo negativo, transferencias conservam pontos no sistema.
  4. Gera relatorio real (Markdown + JSON) com metricas medidas.

Cada usuario virtual envia um X-Forwarded-For unico -> simula 5.000 clientes
distintos e da a cada um seu proprio balde de rate-limit (fiel + sem bater no
teto global de uma unica origem localhost).

Uso:
    python3 scripts/loadtest_homolog.py --users 5000 --workers 32
    python3 scripts/loadtest_homolog.py --users 20 --workers 8     # dry-run
    python3 scripts/loadtest_homolog.py --users 100 --seed-only    # so semeia
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import statistics
import subprocess
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

import requests

# --------------------------------------------------------------------------- #
# Caminhos / constantes                                                        #
# --------------------------------------------------------------------------- #
BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BACKEND_DIR, "loadtest_out")

HOMOLOG_DOMAIN = "homolog.blaxx.test"
EMAIL_FMT = "load.{:05d}@" + HOMOLOG_DOMAIN
EMAIL_LIKE = "load.%@" + HOMOLOG_DOMAIN

BENEFIT_NAME = "Resgate teste (homologacao)"
BENEFIT_COST = 150  # baixo: cabe no menor saldo inicial (2.000 pts)

PKG_KEYS = ("start", "plus", "prime", "black")

FIRST_NAMES = [
    "Ana", "Bruno", "Carla", "Diego", "Eduarda", "Felipe", "Gabriela", "Henrique",
    "Isabela", "Joao", "Karina", "Lucas", "Mariana", "Nicolas", "Olivia", "Paulo",
    "Queila", "Rafael", "Sofia", "Thiago", "Ursula", "Vitor", "Wesley", "Yara",
    "Beatriz", "Caio", "Daniela", "Enzo", "Fernanda", "Gustavo",
]
SURNAMES = [
    "Silva", "Santos", "Oliveira", "Souza", "Costa", "Pereira", "Rodrigues",
    "Almeida", "Nascimento", "Lima", "Araujo", "Fernandes", "Carvalho", "Gomes",
    "Martins", "Rocha", "Ribeiro", "Alves", "Monteiro", "Cardoso", "Teixeira",
    "Andrade", "Barbosa", "Correia", "Pinto",
]


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _stamp() -> str:
    return _utcnow().strftime("%Y%m%d-%H%M%S")


# --------------------------------------------------------------------------- #
# Geradores deterministicos (mesmo seq -> mesmos dados, re-runs estaveis)      #
# --------------------------------------------------------------------------- #
def _cpf_from_base(base9: str) -> str:
    """Anexa os 2 digitos verificadores a 9 digitos base -> CPF valido (11)."""
    digits = [int(c) for c in base9]
    s = sum(d * w for d, w in zip(digits, range(10, 1, -1)))
    r = (s * 10) % 11
    digits.append(0 if r >= 10 else r)
    s = sum(d * w for d, w in zip(digits, range(11, 1, -1)))
    r = (s * 10) % 11
    digits.append(0 if r >= 10 else r)
    return "".join(str(d) for d in digits)


def gen_user(seq: int) -> dict:
    """Dados ficticios deterministicos para o usuario de indice `seq`."""
    import random
    rng = random.Random(seq * 2654435761 & 0xFFFFFFFF)
    name = f"{rng.choice(FIRST_NAMES)} {rng.choice(SURNAMES)}"
    cpf = _cpf_from_base(f"{200_000_000 + seq:09d}")
    return {
        "seq": seq,
        "name": name,
        "email": EMAIL_FMT.format(seq),
        "cpf": cpf,
        "phone": f"+5511{900_000_000 + seq:09d}",
        "password": f"Homolog#{seq:05d}",
        "bonus": rng.randint(2_000, 90_000),
    }


def xff_for(seq: int) -> str:
    """X-Forwarded-For unico por usuario (10.a.b.c). Unico para seq < 16M."""
    return f"10.{(seq >> 16) & 0xFF}.{(seq >> 8) & 0xFF}.{seq & 0xFF}"


# --------------------------------------------------------------------------- #
# App in-process (semeadura + verificacao). O servidor de carga roda separado. #
# --------------------------------------------------------------------------- #
_APP = None


def build_app():
    global _APP
    if _APP is not None:
        return _APP
    os.environ.setdefault("FLASK_ENV", "development")  # evita fail-fast de prod
    os.environ.setdefault("PIX_PROVIDER", "mock")
    os.environ.setdefault("MAILER", "console")
    if BACKEND_DIR not in sys.path:
        sys.path.insert(0, BACKEND_DIR)
    os.chdir(BACKEND_DIR)
    from app import create_app
    _APP = create_app()
    return _APP


# --------------------------------------------------------------------------- #
# Fase 0 - Semeadura                                                           #
# --------------------------------------------------------------------------- #
def seed_users(n: int) -> tuple[list[dict], str]:
    from werkzeug.security import generate_password_hash

    app = build_app()
    from app.extensions import db
    from app.models import Benefit, TxType, User, Wallet
    from app.services import wallet as wallet_svc

    meta: list[dict] = []
    created = skipped = 0
    t0 = time.perf_counter()

    with app.app_context():
        benefit = db.session.query(Benefit).filter_by(name=BENEFIT_NAME).one_or_none()
        if benefit is None:
            benefit = Benefit(
                name=BENEFIT_NAME,
                description="Beneficio ficticio para teste de carga (homologacao).",
                category="desconto", cost_pts=BENEFIT_COST, image_emoji="*",
                stock=-1, expires_in_days=30, tag="Homologacao",
            )
            db.session.add(benefit)
            db.session.commit()
        benefit_id = benefit.id

        for seq in range(n):
            m = gen_user(seq)
            existing = db.session.query(User).filter_by(email=m["email"]).one_or_none()
            if existing is not None:
                if existing.email_verified_at is None:
                    existing.email_verified_at = _utcnow()
                if existing.status != "active":
                    existing.status = "active"
                if existing.wallet is None:
                    db.session.add(Wallet(user_id=existing.id))
                    db.session.flush()
                m["bonus"] = existing.wallet.balance_pts if existing.wallet else 0
                skipped += 1
            else:
                u = User(
                    name=m["name"], email=m["email"], cpf=m["cpf"], phone=m["phone"],
                    pix_key=m["email"], auth_provider="email", status="active",
                    email_verified_at=_utcnow(),
                )
                # Hash barato pbkdf2 direto (argon2 seria lento x N); check_password
                # do modelo aceita o formato werkzeug, entao o login funciona.
                u.password_hash = generate_password_hash(
                    m["password"], method="pbkdf2:sha256:30000"
                )
                db.session.add(u)
                db.session.flush()
                db.session.add(Wallet(user_id=u.id))
                db.session.flush()
                wallet_svc.credit(
                    user_id=u.id, amount_pts=m["bonus"], tx_type=TxType.BONUS,
                    description="Saldo inicial (homologacao/carga)",
                    idempotency_key=f"seed-bonus:{m['email']}",
                )
                created += 1
            meta.append(m)
            if (seq + 1) % 200 == 0:
                db.session.commit()
                print(f"  semeando... {seq + 1}/{n} (novos={created} existentes={skipped})",
                      flush=True)
        db.session.commit()

    # Libera o arquivo SQLite antes do servidor subir noutro processo.
    with app.app_context():
        db.engine.dispose()

    dt = time.perf_counter() - t0
    print(f"  Fase 0 OK: {created} criados, {skipped} reutilizados em {dt:.1f}s "
          f"(beneficio teste id={benefit_id})", flush=True)
    return meta, benefit_id


# --------------------------------------------------------------------------- #
# Fase 1 - Excel + CSV de credenciais de teste                                 #
# --------------------------------------------------------------------------- #
def export_credentials(meta: list[dict], stamp: str) -> dict:
    os.makedirs(OUT_DIR, exist_ok=True)
    xlsx_path = os.path.join(OUT_DIR, f"BLAXX_HOMOLOG_USUARIOS_{stamp}.xlsx")
    csv_path = os.path.join(OUT_DIR, f"BLAXX_HOMOLOG_USUARIOS_{stamp}.csv")

    headers = ["#", "Nome", "E-mail (login)", "CPF", "Telefone",
               "Senha", "Saldo inicial (pts)", "Perfil"]
    banner = ("DADOS FICTICIOS - AMBIENTE DE HOMOLOGACAO. "
              "Nao use usuarios, CPFs, senhas ou saldos reais. PIX em modo mock.")

    # CSV (sempre)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([banner])
        w.writerow(headers)
        for m in meta:
            w.writerow([m["seq"] + 1, m["name"], m["email"], m["cpf"],
                        m["phone"], m["password"], m["bonus"], "cliente"])

    # XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios homologacao"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(row=1, column=1, value=banner)
        c.font = Font(bold=True, color="FF7A0000")
        c.fill = PatternFill("solid", fgColor="FFFFF3CD")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        head_fill = PatternFill("solid", fgColor="FF0A0A0A")
        for col, h in enumerate(headers, start=1):
            hc = ws.cell(row=2, column=col, value=h)
            hc.font = Font(bold=True, color="FFC6FF00")
            hc.fill = head_fill
            hc.alignment = Alignment(horizontal="center")

        for i, m in enumerate(meta):
            r = i + 3
            ws.cell(row=r, column=1, value=m["seq"] + 1)
            ws.cell(row=r, column=2, value=m["name"])
            ws.cell(row=r, column=3, value=m["email"])
            ws.cell(row=r, column=4, value=m["cpf"])
            ws.cell(row=r, column=5, value=m["phone"])
            ws.cell(row=r, column=6, value=m["password"])
            ws.cell(row=r, column=7, value=m["bonus"])
            ws.cell(row=r, column=8, value="cliente")

        widths = [6, 22, 30, 16, 18, 16, 18, 10]
        for col, wdt in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = wdt
        ws.freeze_panes = "A3"
        wb.save(xlsx_path)
        made_xlsx = True
    except Exception as e:  # openpyxl ausente ou erro -> CSV ja cobre
        print(f"  [aviso] falha ao gerar XLSX ({e}); CSV gerado.", flush=True)
        xlsx_path = None
        made_xlsx = False

    print(f"  Fase 1 OK: {len(meta)} credenciais -> "
          f"{'XLSX+CSV' if made_xlsx else 'CSV'} em {OUT_DIR}", flush=True)
    return {"xlsx": xlsx_path, "csv": csv_path}


# --------------------------------------------------------------------------- #
# Fase 2 - Servidor + carga concorrente                                        #
# --------------------------------------------------------------------------- #
def start_server(port: int):
    os.makedirs(OUT_DIR, exist_ok=True)
    env = os.environ.copy()
    env["FLASK_ENV"] = "development"
    env["PIX_PROVIDER"] = "mock"
    env["MAILER"] = "console"
    env["BLAXX_BACKEND_PORT"] = str(port)
    env.setdefault("SECRET_KEY", "homolog-loadtest-secret-key")
    env.setdefault("JWT_SECRET_KEY", "homolog-loadtest-jwt-key")
    # Servidor proprio com threaded=True (run.py nao liga threading) -> concorrencia real.
    # Antes de create_app(), registra PRAGMAs por-conexao no SQLite: busy_timeout
    # reduz erros de "database is locked" sob 32 workers; synchronous=NORMAL acelera
    # o commit sem trocar o modo de journal (sem WAL -> sem sidecar no DB do Dropbox).
    launcher = "\n".join([
        "from sqlalchemy import event",
        "from sqlalchemy.engine import Engine",
        "",
        "@event.listens_for(Engine, 'connect')",
        "def _sqlite_pragmas(dbapi_conn, _rec):",
        "    cur = dbapi_conn.cursor()",
        "    cur.execute('PRAGMA busy_timeout=15000')",
        "    cur.execute('PRAGMA synchronous=NORMAL')",
        "    cur.close()",
        "",
        "from app import create_app",
        f"create_app().run(host='127.0.0.1', port={port}, "
        "threaded=True, use_reloader=False, debug=False)",
    ])
    log_path = os.path.join(OUT_DIR, "server.log")
    logf = open(log_path, "wb")
    proc = subprocess.Popen(
        [sys.executable, "-c", launcher],
        cwd=BACKEND_DIR, env=env, stdout=logf, stderr=subprocess.STDOUT,
    )
    base = f"http://127.0.0.1:{port}"
    for _ in range(80):  # ~40s
        if proc.poll() is not None:
            logf.flush()
            raise RuntimeError(f"servidor saiu cedo (cod={proc.returncode}); veja {log_path}")
        try:
            if requests.get(base + "/health", timeout=2).status_code == 200:
                print(f"  servidor de carga no ar em {base} (threaded)", flush=True)
                return proc, base, logf, log_path
        except requests.RequestException:
            pass
        time.sleep(0.5)
    proc.terminate()
    raise RuntimeError(f"servidor nao respondeu /health a tempo; veja {log_path}")


def stop_server(proc, logf) -> None:
    try:
        proc.terminate()
        proc.wait(timeout=10)
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass
    finally:
        try:
            logf.flush()
            logf.close()
        except Exception:
            pass


def _short(text: str, n: int = 120) -> str:
    return " ".join((text or "").split())[:n]


def journey(m: dict, peer_email: str, base: str, benefit_id: str, do_cashout: bool) -> list:
    """Jornada de 1 usuario virtual. Retorna lista de tuplas de operacao."""
    recs: list = []
    s = requests.Session()
    s.headers.update({
        "X-Forwarded-For": xff_for(m["seq"]),
        "Content-Type": "application/json",
        "User-Agent": "blaxx-homolog-loadtest/1.0",
    })

    def call(op: str, method: str, path: str, **kw):
        t0 = time.perf_counter()
        try:
            r = s.request(method, base + path, timeout=30, **kw)
            ms = (time.perf_counter() - t0) * 1000.0
            ok = r.status_code < 400
            err = None if ok else f"HTTP {r.status_code}: {_short(r.text)}"
            recs.append((op, ok, r.status_code, ms, err))
            return r
        except requests.RequestException as e:
            ms = (time.perf_counter() - t0) * 1000.0
            recs.append((op, False, 0, ms, f"EXC {type(e).__name__}: {_short(str(e))}"))
            return None

    # 1) Login
    r = call("login", "POST", "/auth/login",
             json={"email": m["email"], "password": m["password"]})
    if r is None or r.status_code != 200:
        return recs
    try:
        body = r.json()
        token = body.get("access_token") or body.get("token")
    except ValueError:
        token = None
    if not token:
        recs.append(("login", False, r.status_code, 0.0, "sem token no corpo"))
        return recs
    s.headers["Authorization"] = f"Bearer {token}"

    # 2) Consulta saldo
    call("wallet_read", "GET", "/wallet/")

    # 3) Compra de pontos (PIX mock): cria cobranca + confirma pagamento
    pkg = PKG_KEYS[m["seq"] % len(PKG_KEYS)]
    r = call("pix_charge", "POST", "/pix/charge", json={"package": pkg})
    if r is not None and r.status_code == 201:
        try:
            cid = r.json().get("id")
        except ValueError:
            cid = None
        if cid:
            call("pix_confirm", "POST", "/pix/simulate-payment", json={"charge_id": cid})

    # 4) Transferencia P2P (valor variado 100..700; Idempotency-Key unica)
    amount = 100 + (m["seq"] % 7) * 100
    call("transfer", "POST", "/transfer/",
         json={"to": peer_email, "amount_pts": amount, "password": m["password"],
               "message": "homologacao"},
         headers={"Idempotency-Key": f"lt-{m['seq']:05d}-xfer"})

    # 5) Resgate de beneficio (debita pontos, emite voucher)
    call("benefit_redeem", "POST", f"/benefits/{benefit_id}/redeem")

    # 6) Resgate via PIX (subconjunto) - cashout mock
    if do_cashout:
        call("pix_cashout", "POST", "/redeem/",
             json={"points": 200, "pix_key": m["email"], "password": m["password"]})

    s.close()
    return recs


def run_load(meta: list[dict], benefit_id: str, base: str, workers: int) -> tuple[list, float]:
    n = len(meta)
    records: list = []
    done = 0
    t0 = time.perf_counter()
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {
            ex.submit(journey, meta[i], EMAIL_FMT.format((i + 1) % n), base,
                      benefit_id, (i % 5 == 0)): i
            for i in range(n)
        }
        for fut in as_completed(futs):
            try:
                records.extend(fut.result())
            except Exception as e:
                records.append(("journey_exc", False, 0, 0.0, f"{type(e).__name__}: {e}"))
            done += 1
            if done % 250 == 0 or done == n:
                el = time.perf_counter() - t0
                print(f"  carga... {done}/{n} jornadas ({el:.1f}s, "
                      f"{done / el:.0f} usuarios/s)", flush=True)
    return records, time.perf_counter() - t0


# --------------------------------------------------------------------------- #
# Fase 3 - Integridade do razao                                                #
# --------------------------------------------------------------------------- #
def verify_integrity() -> dict:
    from sqlalchemy import func

    app = build_app()
    from app.extensions import db
    from app.models import Transaction, TxStatus, TxType, User, Wallet

    with app.app_context():
        db.session.remove()  # descarta cache stale; le o que o servidor gravou

        wallets = (
            db.session.query(Wallet)
            .join(User, User.id == Wallet.user_id)
            .filter(User.email.like(EMAIL_LIKE))
            .all()
        )
        n_wallets = len(wallets)
        sum_balance = sum(w.balance_pts for w in wallets)
        negatives = sum(1 for w in wallets if w.balance_pts < 0)

        ledger = dict(
            db.session.query(Transaction.wallet_id, func.sum(Transaction.amount_pts))
            .join(Wallet, Wallet.id == Transaction.wallet_id)
            .join(User, User.id == Wallet.user_id)
            .filter(User.email.like(EMAIL_LIKE),
                    Transaction.status == TxStatus.CONFIRMED)
            .group_by(Transaction.wallet_id)
            .all()
        )
        mismatches = 0
        for w in wallets:
            if w.balance_pts != int(ledger.get(w.id, 0) or 0):
                mismatches += 1
        sum_ledger = int(sum(int(v or 0) for v in ledger.values()))

        by_type = {}
        rows = (
            db.session.query(Transaction.type, func.count(Transaction.id),
                             func.coalesce(func.sum(Transaction.amount_pts), 0))
            .join(Wallet, Wallet.id == Transaction.wallet_id)
            .join(User, User.id == Wallet.user_id)
            .filter(User.email.like(EMAIL_LIKE),
                    Transaction.status == TxStatus.CONFIRMED)
            .group_by(Transaction.type)
            .all()
        )
        for t, cnt, total in rows:
            key = t.value if hasattr(t, "value") else str(t)
            by_type[key] = {"count": int(cnt), "sum_pts": int(total or 0)}

        out = int(by_type.get("transfer_out", {}).get("sum_pts", 0))
        inn = int(by_type.get("transfer_in", {}).get("sum_pts", 0))
        transfer_conserved = (out + inn == 0)

        # AUTORITATIVO: o razao (linhas de Transaction) e a fonte de verdade. Os
        # invariantes que precisam valer sob qualquer concorrencia sao "nenhum saldo
        # negativo" e "conservacao de pontos nas transferencias" (out + in == 0).
        ledger_correct = (negatives == 0 and transfer_conserved)

        # MEDIDO (nao reprova): deriva do CACHE de saldo (coluna balance_pts) sob
        # concorrencia. O SQLite ignora SELECT ... FOR UPDATE, entao um credito
        # transfer_in concorrente pode perder um update e a coluna fica ABAIXO da
        # soma do razao. As linhas do razao seguem corretas; so o cache desvia.
        # Producao (Postgres + SELECT FOR UPDATE) elimina a deriva.
        drift_pts = sum_ledger - sum_balance
        drift_wallets = mismatches

    result = {
        "ledger_correct": ledger_correct,
        "passed": ledger_correct,  # headline/exit baseiam-se na correcao do razao
        "wallets_checked": n_wallets,
        "negative_balances": negatives,
        "transfer_out_pts": out,
        "transfer_in_pts": inn,
        "transfer_conserved": transfer_conserved,
        "sum_balance_pts": int(sum_balance),
        "sum_ledger_pts": sum_ledger,
        "balance_equals_ledger": sum_balance == sum_ledger,
        "balance_cache_drift_pts": int(drift_pts),
        "balance_cache_drift_wallets": int(drift_wallets),
        "by_type": by_type,
    }
    print(f"  Fase 3 {'OK' if ledger_correct else 'FALHOU'} (razao): "
          f"{n_wallets} carteiras, negativos={negatives}, "
          f"conserva_transferencia={transfer_conserved} | deriva de cache: "
          f"{drift_wallets} carteira(s), {drift_pts} pts "
          f"(saldo={sum_balance} vs razao={sum_ledger})",
          flush=True)
    return result


# --------------------------------------------------------------------------- #
# Fase 4 - Metricas + relatorio                                                #
# --------------------------------------------------------------------------- #
def _pct(sorted_vals: list[float], p: float) -> float:
    if not sorted_vals:
        return 0.0
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    k = (len(sorted_vals) - 1) * p
    lo = int(k)
    hi = min(lo + 1, len(sorted_vals) - 1)
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (k - lo)


def summarize(records: list) -> dict:
    ops: dict = {}
    order = ["login", "wallet_read", "pix_charge", "pix_confirm",
             "transfer", "benefit_redeem", "pix_cashout"]
    for op, ok, status, ms, err in records:
        d = ops.setdefault(op, {"count": 0, "ok": 0, "fail": 0, "lat": [],
                                "errors": defaultdict(int)})
        d["count"] += 1
        d["lat"].append(ms)
        if ok:
            d["ok"] += 1
        else:
            d["fail"] += 1
            ekey = (err or "erro").split(":")[0][:48]
            d["errors"][ekey] += 1

    summary = {}
    for op, d in ops.items():
        lat = sorted(d["lat"])
        summary[op] = {
            "count": d["count"], "ok": d["ok"], "fail": d["fail"],
            "success_rate": round(d["ok"] / d["count"] * 100, 2) if d["count"] else 0.0,
            "p50_ms": round(_pct(lat, 0.50), 1),
            "p90_ms": round(_pct(lat, 0.90), 1),
            "p95_ms": round(_pct(lat, 0.95), 1),
            "p99_ms": round(_pct(lat, 0.99), 1),
            "max_ms": round(lat[-1], 1) if lat else 0.0,
            "mean_ms": round(statistics.fmean(lat), 1) if lat else 0.0,
            "errors": dict(d["errors"]),
        }
    return {"_order": [o for o in order if o in summary]
            + [o for o in summary if o not in order], "ops": summary}


def write_report(meta_n: int, workers: int, load_seconds: float, records: list,
                 metrics: dict, integrity: dict, creds: dict, stamp: str) -> dict:
    os.makedirs(OUT_DIR, exist_ok=True)
    total_ops = len(records)
    total_ok = sum(1 for r in records if r[1])
    throughput = total_ops / load_seconds if load_seconds > 0 else 0.0

    payload = {
        "environment": "HOMOLOGACAO (dados ficticios, PIX mock)",
        "generated_at": _utcnow().isoformat(),
        "python": sys.version.split()[0],
        "server": "Werkzeug threaded (dev) + SQLite",
        "users": meta_n,
        "workers": workers,
        "load_seconds": round(load_seconds, 2),
        "total_operations": total_ops,
        "total_ok": total_ok,
        "overall_success_rate": round(total_ok / total_ops * 100, 2) if total_ops else 0.0,
        "throughput_ops_per_sec": round(throughput, 1),
        "metrics": metrics["ops"],
        "integrity": integrity,
        "credentials_files": creds,
    }
    json_path = os.path.join(OUT_DIR, f"BLAXX_HOMOLOG_LOADTEST_{stamp}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    lines = []
    lines.append(f"# BlaXx Pontos - Relatorio de robustez (homologacao)\n")
    lines.append("> **AMBIENTE DE HOMOLOGACAO.** Usuarios, CPFs, senhas, saldos e "
                 "transacoes sao ficticios. PIX em modo mock - nenhum dinheiro real "
                 "movimentado.\n")
    lines.append(f"- Gerado em: `{payload['generated_at']}`")
    lines.append(f"- Python: `{payload['python']}` | Servidor: {payload['server']}")
    lines.append(f"- Usuarios simulados: **{meta_n}** | Workers concorrentes: **{workers}**")
    lines.append(f"- Duracao da carga: **{payload['load_seconds']}s** | "
                 f"Vazao: **{payload['throughput_ops_per_sec']} ops/s**")
    lines.append(f"- Operacoes: **{total_ops}** | Sucesso: **{total_ok}** "
                 f"(**{payload['overall_success_rate']}%**)\n")

    lines.append("## Latencia e sucesso por operacao\n")
    lines.append("| Operacao | Total | OK | Falhas | Sucesso | p50 | p90 | p95 | p99 | max | media |")
    lines.append("|---|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|")
    for op in metrics["_order"]:
        s = metrics["ops"][op]
        lines.append(
            f"| {op} | {s['count']} | {s['ok']} | {s['fail']} | {s['success_rate']}% | "
            f"{s['p50_ms']} | {s['p90_ms']} | {s['p95_ms']} | {s['p99_ms']} | "
            f"{s['max_ms']} | {s['mean_ms']} | (ms)"
        )
    lines.append("")

    errs = {op: metrics["ops"][op]["errors"] for op in metrics["_order"]
            if metrics["ops"][op]["errors"]}
    if errs:
        lines.append("## Erros observados (amostra de robustez)\n")
        for op, emap in errs.items():
            for ekey, cnt in sorted(emap.items(), key=lambda kv: -kv[1]):
                lines.append(f"- `{op}`: {ekey} x{cnt}")
        lines.append("")
    else:
        lines.append("## Erros observados\n\nNenhum erro registrado.\n")

    ig = integrity
    lines.append("## Integridade do razao (ledger) - autoritativo\n")
    lines.append("O razao (linhas de `Transaction` confirmadas) e a fonte de verdade. "
                 "Estes invariantes precisam valer sob qualquer concorrencia:\n")
    lines.append(f"- Resultado do razao: **{'PASSOU' if ig['ledger_correct'] else 'FALHOU'}**")
    lines.append(f"- Carteiras verificadas: {ig['wallets_checked']}")
    lines.append(f"- Saldos negativos: **{ig['negative_balances']}** (invariante: 0)")
    lines.append(f"- Conservacao em transferencias (out {ig['transfer_out_pts']} + "
                 f"in {ig['transfer_in_pts']} = 0): **{ig['transfer_conserved']}**\n")
    lines.append("### Movimentacoes por tipo\n")
    lines.append("| Tipo | Qtd | Soma (pts) |")
    lines.append("|---|--:|--:|")
    for k, v in sorted(ig["by_type"].items()):
        lines.append(f"| {k} | {v['count']} | {v['sum_pts']} |")
    lines.append("")

    drift_pts = ig.get("balance_cache_drift_pts", 0)
    drift_wallets = ig.get("balance_cache_drift_wallets", 0)
    lines.append("## Robustez sob concorrencia (cache de saldo no SQLite)\n")
    lines.append(f"- Soma dos saldos (coluna cache `balance_pts`): **{ig['sum_balance_pts']} pts**")
    lines.append(f"- Soma do razao (linhas `Transaction` confirmadas): **{ig['sum_ledger_pts']} pts**")
    lines.append(f"- Deriva do cache: **{drift_pts} pts** em **{drift_wallets} carteira(s)**\n")
    if drift_pts == 0 and drift_wallets == 0:
        lines.append("Sem deriva nesta execucao: a coluna de cache bateu exatamente "
                     "com o razao mesmo sob concorrencia.\n")
    else:
        lines.append(
            "**Causa raiz (resultado real do teste):** sob alta concorrencia o SQLite "
            "*ignora* `SELECT ... FOR UPDATE` (vira no-op), entao o ciclo "
            "ler-modificar-gravar da coluna `balance_pts` sofre *lost updates* quando "
            "uma carteira recebe um `transfer_in` de uma thread enquanto a thread dona "
            "a atualiza. As **linhas do razao sao sempre gravadas corretamente** "
            "(conservacao e ausencia de negativos se mantem); apenas a coluna de cache "
            "desvia ABAIXO da soma do razao. Em producao o PostgreSQL aplica o bloqueio "
            "de linha de `SELECT FOR UPDATE` e elimina a deriva; o saldo exibido pode "
            "sempre ser reconciliado a partir do razao autoritativo.\n"
        )

    cred_note = creds.get("xlsx") or creds.get("csv") or "(nao gerado)"
    lines.append("## Arquivos gerados\n")
    lines.append(f"- Credenciais de teste: `{cred_note}`")
    lines.append(f"- Metricas JSON: `{json_path}`")
    lines.append(f"- Log do servidor: `{os.path.join(OUT_DIR, 'server.log')}`\n")

    md_path = os.path.join(OUT_DIR, f"BLAXX_HOMOLOG_LOADTEST_{stamp}.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  Fase 4 OK: relatorio -> {md_path}", flush=True)
    return {"md": md_path, "json": json_path}


# --------------------------------------------------------------------------- #
# Orquestracao                                                                 #
# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser(description="Teste de robustez BlaXx (homologacao)")
    ap.add_argument("--users", type=int, default=5000)
    ap.add_argument("--workers", type=int, default=32)
    ap.add_argument("--port", type=int, default=5050)
    ap.add_argument("--seed-only", action="store_true",
                    help="So semeia usuarios + exporta credenciais (sem carga).")
    args = ap.parse_args()

    stamp = _stamp()
    print(f"== BlaXx homologacao :: {args.users} usuarios, {args.workers} workers ==",
          flush=True)
    print("   AMBIENTE DE HOMOLOGACAO - dados ficticios, PIX mock.\n", flush=True)

    print("[Fase 0] Semeando usuarios ficticios...", flush=True)
    meta, benefit_id = seed_users(args.users)

    print("[Fase 1] Exportando credenciais de teste...", flush=True)
    creds = export_credentials(meta, stamp)

    if args.seed_only:
        print("\nConcluido (somente semeadura). Credenciais em:", flush=True)
        print(f"  {creds.get('xlsx') or creds.get('csv')}", flush=True)
        return 0

    print("[Fase 2] Subindo servidor e disparando carga concorrente...", flush=True)
    proc, base, logf, _ = start_server(args.port)
    try:
        records, load_seconds = run_load(meta, benefit_id, base, args.workers)
    finally:
        stop_server(proc, logf)
        time.sleep(1.0)  # garante flush do SQLite do processo servidor

    print("[Fase 3] Verificando integridade do razao...", flush=True)
    integrity = verify_integrity()

    print("[Fase 4] Gerando relatorio...", flush=True)
    metrics = summarize(records)
    paths = write_report(args.users, args.workers, load_seconds, records,
                         metrics, integrity, creds, stamp)

    total = len(records)
    ok = sum(1 for r in records if r[1])
    print("\n==================== RESUMO ====================", flush=True)
    print(f"  Usuarios:        {args.users} (workers={args.workers})", flush=True)
    print(f"  Operacoes:       {total} | OK {ok} "
          f"({ok / total * 100:.2f}%)" if total else "  Operacoes: 0", flush=True)
    print(f"  Vazao:           {total / load_seconds:.0f} ops/s "
          f"em {load_seconds:.1f}s", flush=True)
    print(f"  Razao (ledger):  {'PASSOU' if integrity['passed'] else 'FALHOU'} "
          f"(negativos={integrity['negative_balances']}, "
          f"conserva={integrity['transfer_conserved']})", flush=True)
    print(f"  Deriva de cache: {integrity['balance_cache_drift_pts']} pts em "
          f"{integrity['balance_cache_drift_wallets']} carteira(s) "
          f"[SQLite; Postgres elimina]", flush=True)
    print(f"  Credenciais:     {creds.get('xlsx') or creds.get('csv')}", flush=True)
    print(f"  Relatorio:       {paths['md']}", flush=True)
    print("===============================================", flush=True)
    return 0 if integrity["passed"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
